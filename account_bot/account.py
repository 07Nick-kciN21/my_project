import sqlite3
import openpyxl
import csv
import os
import re
import time
import requests
from bs4 import BeautifulSoup


def insert_object(object, sp):
    workbook = openpyxl.load_workbook('accountbook/book.xlsx')
    thing = re.compile(r'[^0-9]+')
    cost = re.compile(r'[0-9]+')
    t = thing.search(object)
    c = cost.search(object)
    t1 = t.group()
    c1 = int(c.group())
    sheet = workbook.worksheets[0]
    sheet['A1'] = 'Date'
    sheet['B1'] = 'Object'
    sheet['C1'] = 'Cost'
    thing = re.compile(r'[^0-9]+')
    cost = re.compile(r'[0-9]+')
    t = thing.search(object)
    c = cost.search(object)
    d1 = time.strftime("%m/%d", time.localtime())
    t1 = t.group()
    c1 = c.group()
    if(sp == '收入'):
        r = get_rows1()
        sheet['A'+r] = d1
        sheet['B'+r] = t1
        sheet['C'+r] = c1
    if(sp == '支出'):
        r = get_rows6()
        sheet['F'+r] = d1
        sheet['G'+r] = t1
        sheet['H'+r] = c1
    workbook.save('accountbook/book.xlsx')

def delete_object(d_and_o, sp):
    workbook = openpyxl.load_workbook('accountbook/book.xlsx')
    sheet = workbook.worksheets[0]    
    date = re.compile(r'\d\d/\d\d')
    d = date.search(d_and_o)
    d1 = d.group()
    o1 = d_and_o.replace(d1,'')
    r = int(get_rows1())
    list2 = []
    if(sp == '收入'):
        for i in range(2, r+1):
            if(sheet.cell(row=i, column=1).value != d1 or sheet.cell(row=i, column=2).value != o1):
                list1 = []
                list1.append(sheet.cell(row=i, column=1).value)
                list1.append(sheet.cell(row=i, column=2).value)
                list1.append(sheet.cell(row=i, column=3).value)
                list2.append(list1)
        r = 2
        for l in list2:
            print(l)
            sheet['A'+str(r)] = l[0]
            sheet['B'+str(r)] = l[1]
            sheet['C'+str(r)] = l[2]
            r += 1
    r = int(get_rows6())
    if(sp == '支出'):
        for i in range(2, r+1):
            if(sheet.cell(row=i, column=6).value != d1 or sheet.cell(row=i, column=7).value != o1):
                list1 = []
                list1.append(sheet.cell(row=i, column=6).value)
                list1.append(sheet.cell(row=i, column=7).value)
                list1.append(sheet.cell(row=i, column=8).value)
                list2.append(list1)
        r = 2
        for l in list2:
            print(l)
            sheet['F'+str(r)] = l[0]
            sheet['G'+str(r)] = l[1]
            sheet['H'+str(r)] = l[2]
            r += 1
    workbook.save('accountbook/book.xlsx')

def show_object():
    return sort()

def sort():
    workbook = openpyxl.load_workbook('accountbook/book.xlsx')
    sheet = workbook.worksheets[0]
    list1 = []
    cost = 0
    text = ''
    text2= ''
    r = int(get_rows1())
    for i in range(2, r):
        for j in range(1, 4):
            list1.append(sheet.cell(row = i, column = j).value)
        text = '{0} {1:<10} \n金額:{2:^5}\n=======\n'.format(list1[0], list1[1], list1[2])
        cost += int(list1[2])
        text2 += text
        list1 = []
    text2 += '總收入:{}\n///////////////\n'.format(cost)
#==========================================================
    cost = 0
    r = int(get_rows6())
    for i in range(2, r):
        for j in range(6, 9):
            list1.append(sheet.cell(row = i, column = j).value)
        text = '{0} {1:<10} \n金額:{2:^5}\n=======\n'.format(list1[0], list1[1], list1[2])
        cost += int(list1[2])
        text2 += text
        list1 = []
    text2 += '總支出:{}'.format(cost)
    return text2

def create_or_open(name):
    if(os.path.exists(name) != True):
        return True

def get_rows6():
    workbook = openpyxl.load_workbook('accountbook/book.xlsx')
    sheet = workbook.worksheets[0]
    i=1
    while sheet.cell(row=i, column=6).value != None:
        i += 1
    return str(i)

def get_rows1():
    workbook = openpyxl.load_workbook('accountbook/book.xlsx')
    sheet = workbook.worksheets[0]
    i=1
    while sheet.cell(row=i, column=1).value != None:
        i += 1
    return str(i)

